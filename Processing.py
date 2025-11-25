import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import tempfile
import gradio as gr

def normalize_column_name(col_name):
    """Converts column name to lowercase and removes spaces and underscores."""
    return col_name.lower().replace(" ", "").replace("_", "")

def extract_country_code(country_string):
    """Extract the two-letter country code from a country string."""
    if isinstance(country_string, str) and ' - ' in country_string:
        return country_string.split(' - ')[0]
    return country_string

def normalize_value(value):
    """Converts a value to lowercase string, removes spaces and underscores, and handles NaN."""
    if pd.isna(value):
        return None

    # Attempt to convert to a number if it looks like one, then convert to int if possible
    try:
        num_val = float(str(value))
        if num_val == int(num_val):
            value = int(num_val)
        else:
            value = num_val
    except (ValueError, TypeError):
        pass

    return str(value).lower().replace(" ", "").replace("_", "")

def clean_comment(comment):
    """Clean comments by removing 'Discrepancies for Match' lines."""
    if pd.isna(comment) or comment == '':
        return comment
    lines = str(comment).split('\n')
    cleaned_lines = [line for line in lines if not line.strip().startswith('Discrepancies for Match')]
    cleaned = '\n'.join(cleaned_lines).strip()
    return cleaned if cleaned else comment

def process_canf_report(etof_file, rate_card_file, cdp_file, cdp_header_row, cdp_end_column):
    """
    Main processing function for CANF report automation.

    Args:
        etof_file: Uploaded ETOF file
        rate_card_file: Uploaded Rate Card Export file
        cdp_file: Uploaded CDP/SDF file (optional)
        cdp_header_row: Header row number for CDP file (only used if CDP provided)
        cdp_end_column: End column number for CDP file (only used if CDP provided)

    Returns:
        tuple: (output_file_path, status_message)
    """
    try:
        # Validate required files
        if etof_file is None or rate_card_file is None:
            return None, "❌ Error: ETOF file and Rate Card Export file are required."

        # Create output directory - use /content/ in Colab, current dir locally
        import sys
        in_colab = 'google.colab' in sys.modules

        if in_colab:
            # In Colab, save to /content/ so files persist
            output_dir = "/content"
            os.makedirs(output_dir, exist_ok=True)
        else:
            # Locally, use current directory with an output folder
            output_dir = os.path.join(os.getcwd(), "output")
            os.makedirs(output_dir, exist_ok=True)

        # Save uploaded files to processing directory
        etof_path = os.path.join(output_dir, "etof_file.xlsx")
        rate_card_path = os.path.join(output_dir, "rate_card.xlsx")

        # Handle Gradio file uploads
        import shutil
        # Gradio returns file path as string or file object
        if isinstance(etof_file, str):
            shutil.copy(etof_file, etof_path)
        elif hasattr(etof_file, 'name'):
            shutil.copy(etof_file.name, etof_path)
        else:
            shutil.copy(etof_file, etof_path)

        if isinstance(rate_card_file, str):
            shutil.copy(rate_card_file, rate_card_path)
        elif hasattr(rate_card_file, 'name'):
            shutil.copy(rate_card_file.name, rate_card_path)
        else:
            shutil.copy(rate_card_file, rate_card_path)

        # Process CDP file if provided
        df_cdp = None
        if cdp_file is not None:
            # Only use CDP parameters if CDP file is provided
            if cdp_header_row is None or cdp_end_column is None:
                return None, "❌ Error: CDP Header Row and End Column are required when CDP file is provided."

            try:
                cdp_header = int(cdp_header_row)
                cdp_end = int(cdp_end_column)
            except (ValueError, TypeError):
                return None, "❌ Error: CDP Header Row and End Column must be valid numbers."

            # Get file extension
            if isinstance(cdp_file, str):
                file_ext = os.path.splitext(cdp_file)[1]
                cdp_path = os.path.join(output_dir, "cdp_file" + file_ext)
                shutil.copy(cdp_file, cdp_path)
            elif hasattr(cdp_file, 'name'):
                file_ext = os.path.splitext(cdp_file.name)[1]
                cdp_path = os.path.join(output_dir, "cdp_file" + file_ext)
                shutil.copy(cdp_file.name, cdp_path)
            else:
                file_ext = os.path.splitext(str(cdp_file))[1]
                cdp_path = os.path.join(output_dir, "cdp_file" + file_ext)
                shutil.copy(cdp_file, cdp_path)

            # Load CDP file
            file_extension = os.path.splitext(cdp_path)[1].lower()
            if file_extension in ['.xlsx', '.xls']:
                df_cdp = pd.read_excel(cdp_path, header=cdp_header)
            elif file_extension == '.csv':
                df_cdp = pd.read_csv(cdp_path, header=cdp_header)
            else:
                return None, f"❌ Error: Unsupported CDP file format: {file_extension}. Please use .xlsx, .xls, or .csv"

            df_cdp = df_cdp.iloc[:, :cdp_end]
            print("CDP file loaded successfully.")

        # Load ETOF file
        df_etofs = pd.read_excel(etof_path, skiprows=1)

        # Rename duplicate columns
        new_column_names = {
            'Country code': 'Origin Country',
            'Postal code': 'Origin postal code',
            'Airport': 'Origin airport',
            'City': 'Origin city',
            'Country code.1': 'Destination Country',
            'Postal code.1': 'Destination postal code',
            'Airport.1': 'Destination airport',
            'City.1': 'Destination city',
        }
        df_etofs = df_etofs.rename(columns=new_column_names, inplace=False)

        # Select required columns
        df_etofs = df_etofs[['LC #', 'ETOF #', 'Carrier', 'Loading date', 'Equipment type',
                             'Origin Country', 'Origin postal code', 'Origin airport', 'Origin city',
                             'Destination Country', 'Destination postal code', 'Destination airport',
                             'Destination city', 'SERVICE', 'TRANSPORT_MODE', 'SHIPMENT_ID']]

        # Load Rate Card file
        df_rate_card = pd.read_excel(rate_card_path, sheet_name="Rate card", skiprows=2)

        if cdp_file and df_cdp is not None:
            #Rename duplicate columns
            cols = df_cdp.columns.tolist()
            #print(cols)
            new_column_names = {
            'SHAI Reference': 'SHIPMENT_ID',
            'Origin Airport Code': 'Origin airport',
            'Destination Airport Code': 'Destination airport'
              }

            df_cdp = df_cdp.rename(columns=new_column_names, inplace=False)

            df_cdp = df_cdp[['SHIPMENT_ID', 'Origin airport', 'Destination airport']]

            # Check if df_etofs and df_cdp are available
            if df_etofs is not None and df_cdp is not None:
        # Ensure SHIPMENT_ID is string type in both dataframes for consistent merging
        #df_etofs['SHIPMENT_ID'] = df_etofs['SHIPMENT_ID'].astype(str)
        #df_cdp['SHIPMENT_ID'] = df_cdp['SHIPMENT_ID'].astype(str)

        # Merge df_etofs with df_cdp to get the airport information
        # We'll use a left merge to keep all rows from df_etofs
                df_etofs_merged = pd.merge(
                    df_etofs,
                    df_cdp[['SHIPMENT_ID', 'Origin airport', 'Destination airport']],
                    on='SHIPMENT_ID',
                    how='left',
                    suffixes=('', '_cdp') # Suffixes to differentiate columns from df_cdp
                  )

        # Fill NaN values in 'Origin airport' and 'Destination airport' in df_etofs
        # Only update if the df_etofs_merged has the cdp columns
                if 'Origin airport_cdp' in df_etofs_merged.columns:
                    df_etofs['Origin airport'] = df_etofs_merged['Origin airport'].fillna(df_etofs_merged['Origin airport_cdp'])
                if 'Destination airport_cdp' in df_etofs_merged.columns:
                    df_etofs['Destination airport'] = df_etofs_merged['Destination airport'].fillna(df_etofs_merged['Destination airport_cdp'])

        # Drop the temporary merged DataFrame if no longer needed
                del df_etofs_merged

        # Process rate card - find first column index
        first_column_index = None
        if df_rate_card is not None:
            for i, col in enumerate(df_rate_card.columns):
                if "nan" not in str(df_rate_card.iloc[0, i]).lower():
                    first_column_index = i
                    break

        if first_column_index is not None:
            df_rate_card = df_rate_card.iloc[:, :first_column_index]

        if df_rate_card is not None:
            df_rate_card.dropna(subset=[df_rate_card.columns[0]], inplace=True)

        new_columns = df_rate_card.iloc[0].tolist()
        df_rate_card.columns = new_columns
        df_rate_card = df_rate_card.iloc[1:]

        # Get black font columns from rate card
        workbook = openpyxl.load_workbook(rate_card_path)
        sheet = workbook["Rate card"]
        sheet.delete_rows(1, 2)

        first_data_row_index = None
        for row_index, row in enumerate(sheet.iter_rows()):
            first_cell = row[0]
            if first_cell.value is not None:
                first_data_row_index = row_index
                break

        black_font_values = []
        if first_data_row_index is not None:
            first_data_row = list(sheet.iter_rows())[first_data_row_index]
            first_data_values = [cell.value for cell in first_data_row]

            if "Currency" in first_data_values:
                currency_index = first_data_values.index("Currency")
                truncated_data_values = first_data_values[:currency_index]

                for i, value in enumerate(truncated_data_values):
                    cell = first_data_row[i]
                    font_color = "black"
                    if cell.font and cell.font.color:
                        hex_color = cell.font.color.rgb
                        if hex_color is not None and hex_color != '00000000' and hex_color != 'FFFFFFFF':
                            if len(hex_color) == 8:
                                hex_color = hex_color[2:]
                                if len(hex_color) == 6:
                                    r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
                                    if abs(r - g) < 10 and abs(g - b) < 10:
                                        font_color = "grey"
                                    else:
                                        font_color = "other non-black"
                    if font_color == "black":
                        black_font_values.append(value)

        if df_rate_card is not None and black_font_values:
            df_filtered_rate_card = df_rate_card[black_font_values]
        else:
            df_filtered_rate_card = df_rate_card

        # Normalize column names
        df_filtered_rate_card_cols_normalized = [normalize_column_name(col) for col in df_filtered_rate_card.columns]
        df_etofs_cols_normalized = [normalize_column_name(col) for col in df_etofs.columns]

        # Find common columns
        common_columns_normalized = list(set(df_filtered_rate_card_cols_normalized) & set(df_etofs_cols_normalized))

        # Extract country codes
        df_etofs['Origin Country'] = df_etofs['Origin Country'].apply(extract_country_code)
        df_etofs['Destination Country'] = df_etofs['Destination Country'].apply(extract_country_code)

        # Create mappings
        etofs_original_to_normalized = {col: normalize_column_name(col) for col in df_etofs.columns}
        rate_card_original_to_normalized = {col: normalize_column_name(col) for col in df_filtered_rate_card.columns}

        etofs_normalized_to_original = {v: k for k, v in etofs_original_to_normalized.items()}
        rate_card_normalized_to_original = {v: k for k, v in rate_card_original_to_normalized.items()}

        common_etofs_cols_original = [etofs_normalized_to_original[col_norm] for col_norm in common_columns_normalized if col_norm in etofs_normalized_to_original]
        common_rate_card_cols_original = [rate_card_normalized_to_original[col_norm] for col_norm in common_columns_normalized if col_norm in rate_card_normalized_to_original]

        # Pre-calculate unique normalized values
        unique_rc_orig_countries_norm = set(df_filtered_rate_card['Origin Country'].apply(normalize_value).dropna())
        unique_rc_dest_countries_norm = set(df_filtered_rate_card['Destination Country'].apply(normalize_value).dropna())

        unique_rc_orig_cities_norm = set()
        if 'Origin City' in df_filtered_rate_card.columns:
            unique_rc_orig_cities_norm = set(df_filtered_rate_card['Origin City'].apply(normalize_value).dropna())

        unique_rc_dest_cities_norm = set()
        if 'Destination City' in df_filtered_rate_card.columns:
            unique_rc_dest_cities_norm = set(df_filtered_rate_card['Destination City'].apply(normalize_value).dropna())

        # Initialize Comments column
        df_etofs['Comments'] = ''

        # Check for valid from/to columns
        rate_card_cols_normalized = [normalize_column_name(col) for col in df_filtered_rate_card.columns]
        has_valid_from = 'validfrom' in rate_card_cols_normalized
        has_valid_to = 'validto' in rate_card_cols_normalized

        valid_from_col_name = None
        valid_to_col_name = None

        if has_valid_from:
            valid_from_col_name = next((col for col in df_filtered_rate_card.columns if normalize_column_name(col) == 'validfrom'), None)
        if has_valid_to:
            valid_to_col_name = next((col for col in df_filtered_rate_card.columns if normalize_column_name(col) == 'validto'), None)

        # Main matching loop
        for index_etofs, row_etofs in df_etofs.iterrows():
            # Early validation
            etofs_orig_country_norm = normalize_value(row_etofs.get('Origin Country'))
            etofs_dest_country_norm = normalize_value(row_etofs.get('Destination Country'))

            etofs_orig_city_norm = None
            if 'Origin city' in row_etofs:
                etofs_orig_city_norm = normalize_value(row_etofs['Origin city'])

            etofs_dest_city_norm = None
            if 'Destination city' in row_etofs:
                etofs_dest_city_norm = normalize_value(row_etofs['Destination city'])

            if etofs_orig_country_norm is not None and etofs_orig_country_norm not in unique_rc_orig_countries_norm:
                df_etofs.loc[index_etofs, 'Comments'] = f"Origin country: {row_etofs.get('Origin Country', 'N/A')} is not in the rate card."
                continue

            if etofs_dest_country_norm is not None and etofs_dest_country_norm not in unique_rc_dest_countries_norm:
                df_etofs.loc[index_etofs, 'Comments'] = f"Destination country: {row_etofs.get('Destination Country', 'N/A')} is not in the rate card."
                continue

            if etofs_orig_city_norm is not None and etofs_orig_city_norm not in unique_rc_orig_cities_norm:
                df_etofs.loc[index_etofs, 'Comments'] = f"Origin city: {row_etofs.get('Origin city', 'N/A')} is not in the rate card."
                continue

            if etofs_dest_city_norm is not None and etofs_dest_city_norm not in unique_rc_dest_cities_norm:
                df_etofs.loc[index_etofs, 'Comments'] = f"Destination city: {row_etofs.get('Destination city', 'N/A')} is not in the rate card."
                continue

            # Prepare normalized values
            etofs_normalized_values = {
                col_norm: normalize_value(row_etofs.get(common_etofs_cols_original[i]))
                for i, col_norm in enumerate(common_columns_normalized)
                if common_etofs_cols_original[i] in row_etofs
            }

            max_matches = -1
            best_matching_rate_card_rows = []

            # Find best matches
            for index_rate_card, row_rate_card in df_filtered_rate_card.iterrows():
                current_matches = 0

                rate_card_normalized_values = {
                    col_norm: normalize_value(row_rate_card.get(common_rate_card_cols_original[i]))
                    for i, col_norm in enumerate(common_columns_normalized)
                    if common_rate_card_cols_original[i] in row_rate_card
                }

                for col_norm in common_columns_normalized:
                    if col_norm in etofs_normalized_values and col_norm in rate_card_normalized_values:
                        if etofs_normalized_values[col_norm] == rate_card_normalized_values[col_norm]:
                            current_matches += 1

                if current_matches > max_matches:
                    max_matches = current_matches
                    best_matching_rate_card_rows = [{'rate_card_row': row_rate_card.to_dict(), 'discrepancies': []}]
                elif current_matches == max_matches and current_matches > 0:
                    best_matching_rate_card_rows.append({'rate_card_row': row_rate_card.to_dict(), 'discrepancies': []})

            comments_for_current_etofs_row = []

            if len(best_matching_rate_card_rows) > 4:
                df_etofs.loc[index_etofs, 'Comments'] = f"Please recheck the correcntess of shipment data"
                continue

            # Find discrepancies
            for match_idx, best_match_info in enumerate(best_matching_rate_card_rows):
                rate_card_row_dict = best_match_info['rate_card_row']
                discrepancies = []

                # Date validity check
                if has_valid_from and has_valid_to and valid_from_col_name and valid_to_col_name:
                    loading_date = pd.to_datetime(row_etofs.get('Loading date'), errors='coerce')
                    valid_from = pd.to_datetime(rate_card_row_dict.get(valid_from_col_name), errors='coerce')
                    valid_to = pd.to_datetime(rate_card_row_dict.get(valid_to_col_name), errors='coerce')

                    if pd.notna(loading_date) and pd.notna(valid_from) and pd.notna(valid_to):
                        if not (valid_from <= loading_date <= valid_to):
                            discrepancies.append({
                                'column': 'Loading date',
                                'etofs_value': row_etofs.get('Loading date', 'N/A'),
                                'rate_card_value': f"Not within {valid_from.strftime('%Y-%m-%d')} - {valid_to.strftime('%Y-%m-%d')}"
                            })

                for i, col_norm in enumerate(common_columns_normalized):
                    etofs_original_col = common_etofs_cols_original[i]
                    rate_card_original_col = common_rate_card_cols_original[i]

                    etofs_val = row_etofs.get(etofs_original_col)
                    rate_card_val = rate_card_row_dict.get(rate_card_original_col)

                    normalized_etofs_val = normalize_value(etofs_val)
                    normalized_rate_card_val = normalize_value(rate_card_val)

                    if normalized_etofs_val != normalized_rate_card_val:
                        discrepancies.append({
                            'column': etofs_original_col,
                            'etofs_value': etofs_val,
                            'rate_card_value': rate_card_val
                        })
                best_match_info['discrepancies'] = discrepancies

                if discrepancies:
                    comments_for_current_etofs_row.append(f"Discrepancies for Match {match_idx+1}:")
                    for disc in discrepancies:
                        if disc['column'] == 'Loading date' and "Not within" in str(disc['rate_card_value']):
                            comments_for_current_etofs_row.append("  - Shipment date is not within validity range.")
                        else:
                            comments_for_current_etofs_row.append(f"  - {disc['column']}: Shipment value '{disc['etofs_value']}' needs to be changed to '{disc['rate_card_value']}'")

            if comments_for_current_etofs_row:
                df_etofs.loc[index_etofs, 'Comments'] = '\n'.join(comments_for_current_etofs_row)
            else:
                df_etofs.loc[index_etofs, 'Comments'] = 'No discrepancies found for best match.'

        # Export to Excel
        output_filename = os.path.join(output_dir, "Not pre-calculated ETOFs.xlsx")
        df_etofs.to_excel(output_filename, index=False)

        # Create pivot table data
        if 'Comments' in df_etofs.columns:
            cleaned_comments = df_etofs['Comments'].apply(clean_comment)
            pivot_data = cleaned_comments.value_counts().reset_index()
            pivot_data.columns = ['Cause of Not pre-calculation', 'Amount of cases']
            pivot_data = pivot_data.sort_values('Amount of cases', ascending=False)
        else:
            pivot_data = pd.DataFrame({
                'Cause of Not pre-calculation': ['No Comments column found'],
                'Amount of cases': [0]
            })

        # Format Excel file
        workbook = openpyxl.load_workbook(output_filename)
        worksheet = workbook.active
        worksheet.title = "ETOFs Data"

        # Define styles
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Format header row
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border_style

        worksheet.row_dimensions[1].height = 25

        # Format columns
        for col_idx, column in enumerate(worksheet.columns, 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0

            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

            for row_idx, cell in enumerate(column[1:], 2):
                cell.border = border_style
                cell.alignment = left_alignment
                worksheet.row_dimensions[row_idx].height = 20

        worksheet.freeze_panes = 'A2'

        # Create pivot sheet
        pivot_sheet = workbook.create_sheet("Pivot")
        pivot_sheet['A1'] = 'Cause of Not pre-calculation'
        pivot_sheet['B1'] = 'Amount of cases'

        for row_num, (idx, row) in enumerate(pivot_data.iterrows(), start=2):
            pivot_sheet.cell(row=row_num, column=1, value=row['Cause of Not pre-calculation'])
            pivot_sheet.cell(row=row_num, column=2, value=row['Amount of cases'])

        # Format pivot sheet
        for cell in pivot_sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border_style

        pivot_sheet.row_dimensions[1].height = 25

        for col_idx in [1, 2]:
            column_letter = get_column_letter(col_idx)
            max_length = 0

            for row_idx in range(1, len(pivot_data) + 2):
                cell = pivot_sheet.cell(row=row_idx, column=col_idx)
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length

            if col_idx == 1:
                adjusted_width = min(max_length + 2, 80)
            else:
                adjusted_width = min(max_length + 2, 15)
            pivot_sheet.column_dimensions[column_letter].width = adjusted_width

            for row_idx in range(2, len(pivot_data) + 2):
                cell = pivot_sheet.cell(row=row_idx, column=col_idx)
                cell.border = border_style
                if col_idx == 1:
                    cell.alignment = left_alignment
                else:
                    cell.alignment = center_alignment
                pivot_sheet.row_dimensions[row_idx].height = 20

        pivot_sheet.freeze_panes = 'A2'
        workbook.save(output_filename)

        # Prepare status message
        row_count = len(df_etofs)
        status_msg = f"✅ Processing completed successfully!\n\n"
        status_msg += f"📄 Output file: Not pre-calculated ETOFs.xlsx\n"
        status_msg += f"📁 File location: {output_filename}\n"
        status_msg += f"📊 Total rows processed: {row_count}\n"
        status_msg += f"📈 Unique causes: {len(pivot_data)}\n"
        if cdp_file is not None:
            status_msg += f"📋 CDP file processed: Yes\n"
        status_msg += f"\n💡 **Download the file:**\n"
        status_msg += f"   - Use the download button above, or\n"
        status_msg += f"   - Navigate to /content/Not pre-calculated ETOFs.xlsx in Colab file browser and right-click to download"

        # Verify file exists before returning
        if not os.path.exists(output_filename):
            return None, f"❌ Error: Output file was not created successfully."

        # Return the file path - Gradio will handle the download
        return output_filename, status_msg

    except Exception as e:
        import traceback
        error_msg = f"❌ Error processing files: {str(e)}\n\n{traceback.format_exc()}"
        return None, error_msg

def toggle_cdp_inputs(cdp_file):
    """Show/hide CDP inputs based on whether CDP file is provided."""
    if cdp_file is not None:
        return gr.update(visible=True, value=15), gr.update(visible=True, value=33)
    else:
        return gr.update(visible=False), gr.update(visible=False)

# Create Gradio interface
with gr.Blocks(title="CANF Report Automation") as demo:
    gr.Markdown("# 🚀 CANF Report Automation")
    gr.Markdown("Upload your files to process the CANF report. The CDP file is optional.")

    with gr.Row():
        with gr.Column():
            etof_input = gr.File(
                label="ETOF File (Required)",
                file_types=[".xlsx", ".xls"]
            )
            rate_card_input = gr.File(
                label="Rate Card Export File (Required)",
                file_types=[".xlsx", ".xls"]
            )
            cdp_input = gr.File(
                label="CDP/SDF File (Optional)",
                file_types=[".xlsx", ".xls", ".csv"]
            )

        with gr.Column():
            cdp_header_input = gr.Number(
                label="CDP Header Row",
                value=15,
                precision=0,
                visible=False,
                info="Row number where headers start (1-indexed)"
            )
            cdp_end_col_input = gr.Number(
                label="CDP End Column",
                value=33,
                precision=0,
                visible=False,
                info="Last column number to include"
            )

    # Update CDP inputs visibility when CDP file changes
    cdp_input.change(
        fn=toggle_cdp_inputs,
        inputs=[cdp_input],
        outputs=[cdp_header_input, cdp_end_col_input]
    )

    process_btn = gr.Button("🚀 Process Report", variant="primary", size="lg")

    with gr.Row():
        with gr.Column():
            output_file = gr.File(label="📥 Download Output File")
        with gr.Column():
            status_output = gr.Textbox(label="📊 Status", lines=12, interactive=False)

    def process_and_store(etof_file, rate_card_file, cdp_file, cdp_header_row, cdp_end_column):
        """Process files and return the output."""
        file_path, status = process_canf_report(etof_file, rate_card_file, cdp_file, cdp_header_row, cdp_end_column)
        if file_path:
            return file_path, status
        else:
            return None, status

    process_btn.click(
        fn=process_and_store,
        inputs=[etof_input, rate_card_input, cdp_input, cdp_header_input, cdp_end_col_input],
        outputs=[output_file, status_output]
    )

    gr.Markdown("### 📖 Instructions")
    gr.Markdown("""
    1. **Upload ETOF file** (required) - Excel file containing ETOF data
    2. **Upload Rate Card Export file** (required) - Excel file with rate card data
    3. **Upload CDP/SDF file** (optional) - If provided, you'll need to specify:
       - Header Row: The row number where column headers start
       - End Column: The last column number to include
    4. Click **"Process Report"** button
    5. Download the output file:

    **Download Options:**
    - **Gradio Download Button**: Click the download button in the "Download Output File" section above
    - **Colab File Browser**: Navigate to `/content/Not pre-calculated ETOFs.xlsx` in the Colab file browser (left sidebar) and right-click to download
    - **Programmatic Download**: Run this in a new cell after processing:
      ```python
      from google.colab import files
      files.download("/content/Not pre-calculated ETOFs.xlsx")
      ```

    **Note:** The CDP Header Row and End Column inputs will only appear when a CDP file is uploaded.
    """)

# Launch Gradio interface
if __name__ == "__main__":
    # For Google Colab, we need to use share=True to get a public URL
    # This is required for Colab to access the interface properly
    import sys

    # Check if running in Colab
    in_colab = 'google.colab' in sys.modules

    if in_colab:
        # In Colab, you can use share=False for local access or share=True for public URL
        # share=False: Access via Colab's proxy (more secure, but download button may have issues)
        # share=True: Public URL (easier downloads, but less secure)
        use_share = False  # Change to False if you prefer local access

        if use_share:
            print("🚀 Launching Gradio interface for Google Colab (public URL)...")
            print("💡 Using share=True - you'll get a public URL for easy access")
            demo.launch(share=True, debug=False, show_error=True)
        else:
            print("🚀 Launching Gradio interface for Google Colab (local access)...")
            print("💡 Using share=False - access via Colab's proxy")
            print("💡 Download options: Use 'Download via Colab Files' button or Colab file browser")
            demo.launch(share=False, server_name="0.0.0.0", debug=False, show_error=True)
    else:
        # For local execution
        print("🚀 Launching Gradio interface locally...")
        print("💡 Output files will be saved to: ./output/")
        demo.launch(server_name="127.0.0.1", server_port=7860, share=False)
